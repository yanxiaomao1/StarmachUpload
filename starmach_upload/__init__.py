'''data importer'''

import logging
import Database
from eventmanager import Evt
from data_import import DataImporter
import re


from openpyxl import load_workbook
from io import BytesIO

logger = logging.getLogger(__name__)


def import_pilot(importer_class, rhapi, source, args):
    if not source:
        return False
    try:
        excel_file = BytesIO(source)
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
    except Exception:
        return False
    try:
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index != 0 and row[0] != None and row[0] != '':
                pilot = {
                    "name": row[0],
                    "callsign": row[1],
                }
                new_pilot = rhapi.db.pilot_add(name=pilot["name"], callsign=pilot["callsign"])
                rhapi.db.pilot_alter(pilot_id=new_pilot.id, attributes={
                    'uuid': row[0]
                })
    except Exception:
        return False
    return True

def import_heat(importer_class, rhapi, source, args):
    if not source:
        return False
    try:
        excel_file = BytesIO(source)
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
    except Exception:
        return False
        
    try:
        group = -1
        slot_id = 0
        for index, row in enumerate(sheet.iter_rows(values_only=True)):

            if index == 0:
                name = row[0]
                raceclass = rhapi.db.raceclass_add(name=name)
                class_id = raceclass.id
            elif index != 1:
                group_name = row[0]
                pilot_id = row[1]
                # pilot_name = row[2]
                # channel = row[3]
                led_color = row[4]

                if group_name != None and (group == -1 or group != group_name):
                    group = group_name
                    heat = rhapi.db.heat_add(name=group_name, raceclass=class_id, auto_frequency=False)
                    heat_id = heat.id
                    slot_id = 0
                else:
                    slot_id = slot_id + 1
                    
                pilot = Database.Pilot.query.filter_by(name = str(pilot_id)).first()
                pilot.color = get_led_color(led_color)
                heats = Database.HeatNode.query.filter_by(heat_id = heat_id, node_index = slot_id).first()
                heats.pilot_id = pilot.id
                Database.DB_session.commit()

    except Exception:
        return False
            
    return True

def get_led_color(color):
    
    new_color =  '#FFFFFF'

    if is_hex(color):
        new_color = color
    else:
        if color == '红':
            new_color = '#FF0000'
        elif color == '黄':
            new_color = '#FFFF00'
        elif color == '蓝':
            new_color = '#0000FF'
        elif color == '绿':
            new_color = '#008000'
        elif color == '青':
            new_color = '#00FFFF'
        elif color == '品红':
            new_color = '#FF00FF'

    return new_color
    

def is_hex(s):
    if s.startswith("0x") or s.startswith("#"):
        s = s[2:]
    return re.fullmatch(r"[0-9A-Fa-f]+", s) is not None
    

def register_handlers(args):
    for importer in [
        DataImporter(
            '导入飞手',
            import_pilot,
            None,
            None,
        ),
        DataImporter(
            '导入分组',
            import_heat,
            None,
            None,
        ),
    ]:
        args['register_fn'](importer)

def initialize(rhapi):
    rhapi.events.on(Evt.DATA_IMPORT_INITIALIZE, register_handlers)
